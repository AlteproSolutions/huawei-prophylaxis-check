from ciscoconfparse import CiscoConfParse
from openpyxl import Workbook
from scrapli import Scrapli
from rich import inspect
from rich import print as rprint
import os
import sys
import re
from dotenv import load_dotenv
from scrapli.exceptions import ScrapliException
import concurrent.futures
import pandas as pd
from tqdm.contrib.concurrent import thread_map
from netmiko.ssh_autodetect import SSHDetect
import logging


# Initialize logging
logging.basicConfig(filename='magic.log', filemode='w', format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)

def autodetect_device(host, username, password):
    device = {
    'device_type': 'autodetect',
    'ip': host,
    'username': username,
    'password': password,
}
    
    guesser = SSHDetect(**device)
    best_match = guesser.autodetect()
    print(best_match) # Name of the best device_type to use further
    print(guesser.potential_matches) # Dictionary of the whole matching result


def create_inventory(hosts, username, password):

    inventory = []


    for host in hosts:
        inventory.append({
            "host": host.strip(),
            "auth_username": username,
            "auth_password": password,
            "auth_strict_key": False,
            "platform": "huawei_vrp",
            "transport": "ssh2"
        })

    return inventory


def connect_to_device(device):
    try:
        conn = Scrapli(**device)
        conn.open()
        return conn
    
    except ScrapliException as e:
        logging.error(f"ScrapliException: {e}")
        


def get_config(connected_device):
    try:
        return connected_device.send_command("display current-configuration").result

    except ScrapliException as e:
        logging.error(f"ScrapliException: {e}")
        


def get_hostname(config):
    for line in config.splitlines():
        if line.startswith("sysname"):
            hostname = line.split()[1]
            return hostname
    return "HOSTNAME_NOT_FOUND"


def perform_live_checks(device_connection):

    live_checks = {}

    try:
        # Check no BPDU error-down
        response = device_connection.send_command("display error-down recovery")
        output = response.result
        live_checks["check_no_bpdu_error_down"] = "Info: No error-down interface exists." in output

        # Check NTP status
        response = device_connection.send_command("display ntp status | include clock status")
        output = response.result
        live_checks["check_ntp_status_synchronized"] = "clock status: synchronized" in output

        # Check HTTP status
        response = device_connection.send_command("display http server")
        output = response.result
        live_checks["check_http_status_disabled"] = "HTTP Server Status              : disabled" and "HTTP Secure-server Status       : disabled" in output

    except ScrapliException as e:
        logging.error(f"ScrapliException: {e}")

    return live_checks


def check_config(parsed_config, global_lines_to_check):

    result = {}

    for line in global_lines_to_check:
        if parsed_config.find_objects(line, exactmatch=True):
            result[line] = True
        else:
            result[line] = False

    #check STP Mode
    stp_mode_line = parsed_config.find_lines(r"^spanning-tree mode")
    if stp_mode_line:
        stp_mode = stp_mode_line[0].split()[-1]
        result["STP Mode"] = stp_mode

    return result


def check_interfaces_config(parsed_config, interface_filter, interfaces_lines_to_check):

    interfaces = parsed_config.find_objects_w_child(parentspec=r'^interface', childspec=interface_filter)

    interface_check_results = {}

    for interface in interfaces:
        checks = {}
    
    # Check each line in interfaces_lines_to_check
        for line in interfaces_lines_to_check:
            checks[line] = any([cfg_line.text.strip() == line for cfg_line in interface.children])
        
        interface_check_results[interface.text.strip()] = checks

    return interface_check_results


def save_to_excel(data, output_file):

    print("\nSaving data to excel file ...")

    # Create a new workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Outputs"

    # Set column headers
    headers = ['Hostname', 'IP address']
    for device in data:
        for ip, info in device.items():
            for check_name in info['check_results']:
                if check_name not in headers:
                    headers.append(check_name)

    for col_num, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col_num).value = header

    # Add data to the worksheet
    for row_num, device in enumerate(data, 2):
        for ip, info in device.items():
            ws1.cell(row=row_num, column=1).value = info['hostname']
            ws1.cell(row=row_num, column=2).value = ip
            for col_num, header in enumerate(headers[2:], 3):
                ws1.cell(row=row_num, column=col_num).value = info['check_results'].get(header, None)

    # Create a new sheet for interfaces_check
    ws2 = wb.create_sheet("Interfaces Check")

    # Set column headers for the new sheet
    headers = ['IP address', 'Hostname', 'Interface']
    for device in data:
        for ip, info in device.items():
            for iface, iface_info in info['interfaces_check'].items():
                for check_name in iface_info:
                    if check_name not in headers:
                        headers.append(check_name)

    for col_num, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col_num).value = header

    # Add data to the new sheet
    row_num = 2
    for device in data:
        for ip, info in device.items():
            for interface, interface_info in info['interfaces_check'].items():
                ws2.cell(row=row_num, column=1).value = ip
                ws2.cell(row=row_num, column=2).value = info['hostname']
                ws2.cell(row=row_num, column=3).value = interface
                for col_num, header in enumerate(headers[3:], 4):
                    ws2.cell(row=row_num, column=col_num).value = interface_info.get(header, None)
                row_num += 1

    # Save the workbook to a file
    wb.save(output_file)


def process_device(device, global_lines_to_check, interfaces_lines_to_check, interfaces_filter, commands_to_get_output):

    try:
        conn = connect_to_device(device)
        device_config = get_config(conn)
        hostname = get_hostname(device_config)
        ip = device.get("host")
        device_path = f"{absolute_path}/output/{hostname}_" + f'({ip})'
        confparse = CiscoConfParse(device_config.splitlines())

        if not os.path.exists(device_path): os.mkdir(device_path)
        
        with open(f'{device_path}/config.txt', mode="w") as device_config_file:
            device_config_file.write(device_config)
            print(f"\nConfig saved for {device}")

    except Exception as e:
        logging.error(f"Failed when getting config from {device}, error: {e}")
        print(f"\nFailed when getting config from {device}, error: {e}")
        return

    #Get commands and save the output
    for command in commands_to_get_output:
        try:
            response = conn.send_command(command)

            with open(f'{device_path}/{command.strip().replace(" ", "_")}-{hostname}_({ip}).txt', mode="w") as commandfile:
                commandfile.write(response.result)

        except ScrapliException as e:
            logging.error(f"ScrapliException: {e}")
            print("Device " + hostname + f' ({ip}) - failed when getting ' + command)


    #Perform config and live checks
    print("Performing config and live checks for device", hostname)

    live_checks = perform_live_checks(conn)
    config_checks = check_config(confparse, global_lines_to_check)
    interfaces_check = check_interfaces_config(confparse, interfaces_filter, interfaces_lines_to_check)

    conn.close()

    checks_results = {}
    checks_results.update(live_checks)
    checks_results.update(config_checks)

    device_data = {
        ip: {
            'hostname': hostname,
            'config': device_config,
            'check_results': checks_results,
            'interfaces_check': interfaces_check,
        }
    }

    return device_data


def process_device_with_args(args):
        return process_device(*args)


if __name__ == "__main__":

    #Load environment variables
    absolute_path = os.path.dirname(os.path.realpath(__file__))
    load_dotenv(absolute_path + "/.env")

    #Load config
    UseNetmikoAutodetect = False



    username = os.environ.get("USER")
    password = os.environ.get("PASSWORD")    # Load Environment Variables
    
    if (username or password) is None:
        print('U need to fill the .env file in root directory of the script and add USER = "YOURUSER" and PASSWORD = "YOURPASSWORD"')
        sys.exit(1)





    # Create output folder
    if not os.path.exists(f'{absolute_path}/output'): os.mkdir(f'{absolute_path}/output')

    # Read hosts
    with open(f'{absolute_path}/config/hosts.txt', mode="r") as hostsFile:
        hosts = hostsFile.readlines()

    # Read commands
    with open(f'{absolute_path}/config/commands.txt', mode="r") as f:
        commands_to_get_ouput = f.read().splitlines()

    # Read lines to check
    with open(f'{absolute_path}/config/global_lines_to_check.txt', mode="r") as f:
        global_lines_to_check = f.read().splitlines()

    # Read interface lines to check
    with open(f'{absolute_path}/config/interface_lines_to_check.txt', mode="r") as f:
        interface_lines_to_check = f.read().splitlines()

    # Regex filter for interfaces
    with open(f'{absolute_path}/config/regex_interfaces_filter.txt', mode="r") as f:
        interfaces_filter = f.read().strip()
    

    print("Commands: ", commands_to_get_ouput)
    print("Interfaces filter: ", interfaces_filter)
    print("Devices : ", list(map(lambda x:x.strip(),hosts)))



    inventory = create_inventory(hosts, username, password)

    devices_data = []

    #with concurrent.futures.ThreadPoolExecutor() as executor:
     #   futures = [executor.submit(process_device, device, global_lines_to_check, commands_to_get_ouput) for device in tqdm(inventory, desc="Processing devices ...", unit="device")]
     #   concurrent.futures.wait(futures)

     #   for future in concurrent.futures.as_completed(futures):
     #       devices_data.append(future.result())

    devices_data = thread_map(
        process_device_with_args,
        [(device, global_lines_to_check, interface_lines_to_check, interfaces_filter, commands_to_get_ouput) for device in inventory],
        desc="Processing devices",
        unit="device",
    )

    # Filter out None values from the devices_data list
    devices_data = [device for device in devices_data if device is not None]
    
    save_to_excel(devices_data, absolute_path + "/output/commands_check.xlsx")

    print("Script has finished")

